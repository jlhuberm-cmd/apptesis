"""Exportador de reportes del dashboard a Excel y PDF.

Toma el MISMO contexto que construye `SupabaseDashboardService.build_context`
(el que ya alimenta el tablero en pantalla) y lo materializa en un archivo
descargable. De este modo el reporte refleja exactamente lo que el usuario ve,
incluidos los filtros (genero, rango de edad, competencia) aplicados.

No modifica ningun dato: es una capa de presentacion adicional, de solo lectura.

Formatos:
- Excel (.xlsx) via openpyxl: hojas Resumen, Estadistica y Distribucion.
- PDF (.pdf) via reportlab: documento de una o dos paginas con las mismas tablas.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    Image as RLImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Colores institucionales UTPL.
_UTPL_AZUL = "003B71"
_UTPL_NARANJA = "F39C12"
_GRIS_CLARO = "F2F4F7"

_TITULO = "Reporte de Competencias Digitales - DigComp 2.2"
_SUBTITULO = "Area 4: Seguridad - UTPL"

# Orden y titulos de los graficos capturados desde el tablero (Plotly).
_GRAFICOS = (
    ("radar", "Perfil de competencias (radar)"),
    ("bar", "Media por competencia"),
    ("dist", "Distribucion por nivel"),
)


def _describe_filtros(context: dict) -> str:
    """Texto legible de los filtros aplicados al tablero."""
    f = context.get("current_filters", {}) or {}
    partes = []
    if f.get("respondent_gender"):
        partes.append(f"Genero: {f['respondent_gender']}")
    if f.get("respondent_age_range"):
        partes.append(f"Rango de edad: {f['respondent_age_range']}")
    selected = context.get("selected_code")
    names = context.get("competency_names", {})
    if selected:
        partes.append(f"Competencia (distribucion): {selected} - {names.get(selected, '')}")
    return " | ".join(partes) if partes else "Sin filtros (todos los encuestados)"


# ====================================================================== #
# EXCEL
# ====================================================================== #
def to_excel(context: dict, images: dict[str, bytes] | None = None) -> bytes:
    """Construye el libro Excel del dashboard y devuelve sus bytes.

    `images` (opcional): {clave: png_bytes} con los graficos capturados del
    tablero (radar, bar, dist). Si se entregan, se agregan en una hoja Graficos.
    """
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor=_UTPL_AZUL)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, color=_UTPL_AZUL, size=14)
    sub_font = Font(italic=True, color="555555", size=10)
    center = Alignment(horizontal="center", vertical="center")

    def _style_header(ws, row: int, n_cols: int) -> None:
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    def _autosize(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --- Hoja 1: Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = _TITULO
    ws["A1"].font = title_font
    ws["A2"] = _SUBTITULO
    ws["A2"].font = sub_font
    ws["A3"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = sub_font
    ws["A4"] = f"Filtros: {_describe_filtros(context)}"
    ws["A4"].font = sub_font
    ws["A5"] = f"Muestra: {context.get('sample_size', 0)} encuestados"
    ws["A5"].font = sub_font

    stats = context.get("stats")
    overall = getattr(stats, "overall_mean", 0.0) if stats else 0.0
    ws["A6"] = f"Media global (autoevaluacion, escala 1-4): {overall}"
    ws["A6"].font = Font(bold=True, color=_UTPL_NARANJA, size=11)

    encabezados = ["Codigo", "Competencia", "Media Autoeval. (1-4)",
                   "Media Conocimiento (1-4)", "Nivel"]
    hrow = 8
    for col, txt in enumerate(encabezados, start=1):
        ws.cell(row=hrow, column=col, value=txt)
    _style_header(ws, hrow, len(encabezados))

    r = hrow + 1
    for card in context.get("cards", []):
        ws.cell(row=r, column=1, value=card["code"])
        ws.cell(row=r, column=2, value=card["name"])
        ws.cell(row=r, column=3, value=card["mean"])
        ws.cell(row=r, column=4, value=card["conoc"])
        ws.cell(row=r, column=5, value=card["category"])
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=_GRIS_CLARO)
        r += 1
    _autosize(ws, [12, 42, 22, 24, 14])

    # --- Hoja 2: Estadistica descriptiva ---
    ws2 = wb.create_sheet("Estadistica")
    ws2["A1"] = "Estadistica descriptiva por competencia (autoevaluacion 1-4)"
    ws2["A1"].font = title_font
    cols = ["Codigo", "Competencia", "Media", "Mediana", "Moda", "Desv. Est.",
            "Varianza", "Asimetria", "Curtosis", "Min", "Max", "N",
            "P25", "P50", "P75"]
    hrow2 = 3
    for col, txt in enumerate(cols, start=1):
        ws2.cell(row=hrow2, column=col, value=txt)
    _style_header(ws2, hrow2, len(cols))

    r = hrow2 + 1
    comps = getattr(stats, "competencies", []) if stats else []
    for s in comps:
        valores = [s.competency_code, s.competency_name, s.mean, s.median, s.mode,
                   s.std_deviation, s.variance, s.skewness, s.kurtosis, s.min_value,
                   s.max_value, s.count, s.percentile_25, s.percentile_50,
                   s.percentile_75]
        for col, v in enumerate(valores, start=1):
            ws2.cell(row=r, column=col, value=v)
        if r % 2 == 0:
            for c in range(1, len(cols) + 1):
                ws2.cell(row=r, column=c).fill = PatternFill("solid", fgColor=_GRIS_CLARO)
        r += 1
    _autosize(ws2, [10, 38, 9, 9, 9, 11, 10, 10, 10, 7, 7, 6, 8, 8, 8])

    # --- Hoja 3: Distribucion por nivel ---
    ws3 = wb.create_sheet("Distribucion")
    dist = context.get("dist")
    nombre = getattr(dist, "competency_name", "") if dist else ""
    ws3["A1"] = f"Distribucion por nivel - {nombre}"
    ws3["A1"].font = title_font
    ws3["A2"] = f"Linea de media: {getattr(dist, 'mean_line', 0.0) if dist else 0.0}"
    ws3["A2"].font = sub_font
    enc3 = ["Nivel", "Frecuencia"]
    hrow3 = 4
    for col, txt in enumerate(enc3, start=1):
        ws3.cell(row=hrow3, column=col, value=txt)
    _style_header(ws3, hrow3, len(enc3))
    r = hrow3 + 1
    if dist:
        for nivel, freq in zip(dist.bins, dist.frequencies):
            ws3.cell(row=r, column=1, value=nivel)
            ws3.cell(row=r, column=2, value=freq)
            r += 1
    _autosize(ws3, [20, 14])

    # --- Hoja 4: Graficos del tablero (capturados desde Plotly) ---
    if images:
        ws4 = wb.create_sheet("Graficos")
        ws4["A1"] = "Graficos del tablero (segun filtros aplicados)"
        ws4["A1"].font = title_font
        ws4.column_dimensions["A"].width = 75
        _img_refs: list[BytesIO] = []  # mantener vivas las referencias hasta save()
        fila = 3
        for clave, titulo in _GRAFICOS:
            data = images.get(clave)
            if not data:
                continue
            ws4.cell(row=fila, column=1, value=titulo).font = Font(
                bold=True, color=_UTPL_AZUL, size=11)
            bio = BytesIO(data)
            xlimg = XLImage(bio)
            xlimg.width, xlimg.height = 480, 270  # 16:9
            ws4.add_image(xlimg, f"A{fila + 1}")
            _img_refs.append(bio)
            fila += 17  # espacio suficiente para la imagen

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ====================================================================== #
# PDF
# ====================================================================== #
def to_pdf(context: dict, images: dict[str, bytes] | None = None) -> bytes:
    """Construye el reporte PDF del dashboard y devuelve sus bytes.

    `images` (opcional): {clave: png_bytes} con los graficos capturados del
    tablero (radar, bar, dist). Si se entregan, se agregan en una pagina final.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        title=_TITULO,
    )

    azul = colors.HexColor(f"#{_UTPL_AZUL}")
    naranja = colors.HexColor(f"#{_UTPL_NARANJA}")
    gris = colors.HexColor(f"#{_GRIS_CLARO}")

    styles = getSampleStyleSheet()
    h_title = ParagraphStyle("titulo", parent=styles["Title"], textColor=azul,
                             fontSize=18, spaceAfter=2)
    h_sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.HexColor("#555555"),
                           fontSize=10, spaceAfter=1)
    h_section = ParagraphStyle("seccion", parent=styles["Heading2"], textColor=azul,
                               fontSize=13, spaceBefore=12, spaceAfter=6)
    h_metric = ParagraphStyle("metric", parent=styles["Normal"], textColor=naranja,
                              fontSize=11, spaceAfter=8)

    stats = context.get("stats")
    overall = getattr(stats, "overall_mean", 0.0) if stats else 0.0

    elementos = []
    elementos.append(Paragraph(_TITULO, h_title))
    elementos.append(Paragraph(_SUBTITULO, h_sub))
    elementos.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", h_sub))
    elementos.append(Paragraph(f"Filtros: {_describe_filtros(context)}", h_sub))
    elementos.append(Paragraph(f"Muestra: {context.get('sample_size', 0)} encuestados", h_sub))
    elementos.append(Spacer(1, 6))
    elementos.append(Paragraph(
        f"Media global (autoevaluacion, escala 1-4): {overall}", h_metric))

    def _tabla(data, col_widths=None, header_bg=azul):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, gris]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        t.setStyle(TableStyle(estilo))
        return t

    # --- Resumen por competencia ---
    elementos.append(Paragraph("Resumen por competencia", h_section))
    resumen = [["Codigo", "Competencia", "Media Autoeval.", "Media Conoc.", "Nivel"]]
    for card in context.get("cards", []):
        resumen.append([card["code"], card["name"], str(card["mean"]),
                        str(card["conoc"]), card["category"]])
    elementos.append(_tabla(resumen, col_widths=[2.2 * cm, 10 * cm, 3.5 * cm, 3.5 * cm, 3 * cm]))

    # --- Estadistica descriptiva ---
    elementos.append(Paragraph("Estadistica descriptiva (autoevaluacion 1-4)", h_section))
    enc = ["Cod.", "Media", "Mediana", "Moda", "Desv.", "Var.", "Asim.",
           "Curt.", "Min", "Max", "N", "P25", "P50", "P75"]
    cuerpo = [enc]
    comps = getattr(stats, "competencies", []) if stats else []
    for s in comps:
        cuerpo.append([
            s.competency_code, s.mean, s.median,
            "-" if s.mode is None else s.mode,
            s.std_deviation, s.variance, s.skewness, s.kurtosis,
            s.min_value, s.max_value, s.count,
            s.percentile_25, s.percentile_50, s.percentile_75,
        ])
    elementos.append(_tabla(cuerpo, header_bg=naranja))

    # --- Distribucion por nivel ---
    dist = context.get("dist")
    if dist:
        elementos.append(Paragraph(
            f"Distribucion por nivel - {dist.competency_name}", h_section))
        dist_data = [["Nivel", "Frecuencia"]]
        for nivel, freq in zip(dist.bins, dist.frequencies):
            dist_data.append([nivel, str(freq)])
        elementos.append(_tabla(dist_data, col_widths=[6 * cm, 4 * cm]))

    # --- Graficos del tablero (capturados desde Plotly) ---
    if images and any(images.get(k) for k, _ in _GRAFICOS):
        elementos.append(PageBreak())
        elementos.append(Paragraph("Graficos del tablero", h_section))
        for clave, titulo in _GRAFICOS:
            data = images.get(clave)
            if not data:
                continue
            elementos.append(Paragraph(titulo, h_sub))
            elementos.append(RLImage(BytesIO(data), width=14 * cm, height=7.875 * cm))
            elementos.append(Spacer(1, 10))

    doc.build(elementos)
    return buffer.getvalue()
